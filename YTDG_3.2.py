import os
import json
import threading
import subprocess
import yt_dlp
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import load_workbook, Workbook

EXCEL_FILE = 'videos.xlsx'

# ----------------- YouTube 搜索 -----------------
def search_youtube(query, search_pool_size):
    command = [
        'yt-dlp',
        f'ytsearch{search_pool_size}:{query}',
        '--dump-json'
    ]
    result = subprocess.run(command, capture_output=True, text=True)
    videos = []
    for line in result.stdout.strip().split('\n'):
        video = json.loads(line)
        videos.append(video)
    return videos

# ----------------- Excel 读写 -----------------
def load_existing_links(file_path):
    if not os.path.exists(file_path):
        return set()
    wb = load_workbook(file_path)
    ws = wb.active
    return set(row[0] for row in ws.iter_rows(values_only=True) if row[0])

def append_links_to_excel(file_path, new_links):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Video Link'])

    for link in new_links:
        ws.append([link])
    wb.save(file_path)

# ----------------- 视频下载 -----------------
def download_video(video_url, quality, output_dir='downloads', progress_callback=None):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    format_map = {
        '480p': 'bestvideo[ext=mp4][height<=480]+bestaudio[ext=m4a]/best[ext=mp4][height<=480]',
        '720p': 'bestvideo[ext=mp4][height<=720]+bestaudio[ext=m4a]/best[ext=mp4][height<=720]',
        '1080p': 'bestvideo[ext=mp4][height<=1080]+bestaudio[ext=m4a]/best[ext=mp4][height<=1080]',
    }

    ydl_opts = {
        'outtmpl': os.path.join(output_dir, '%(title)s.%(ext)s'),
        'quiet': True,
        'no_warnings': True,
        'format': format_map.get(quality, format_map['480p']),
        'merge_output_format': 'mp4',
        'progress_hooks': [progress_callback] if progress_callback else [],
    }

    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        try:
            ydl.download([video_url])
            return True
        except Exception as e:
            print(f"下载失败: {video_url} 原因: {e}")
            return False

# ----------------- 下载主线程 -----------------
def download_thread(keyword, desired_download_count, max_duration_minutes, quality):
    max_duration_seconds = max_duration_minutes * 60
    search_pool_size = desired_download_count * 5

    log_text.delete('1.0', tk.END)
    log_text.insert(tk.END, f"\n开始搜索: {keyword}  目标: {desired_download_count} 时长限制: {max_duration_minutes} 分钟 画质: {quality}\n")
    root.update()

    results = search_youtube(keyword, search_pool_size)
    log_text.insert(tk.END, f"已获取 {len(results)} 条搜索结果\n")
    root.update()

    existing_links = load_existing_links(EXCEL_FILE)
    new_links = []
    downloaded_count = 0

    for video in results:
        if downloaded_count >= desired_download_count:
            break

        link = video['webpage_url']
        duration_sec = video.get('duration', 0)
        title = video.get('title', 'Unknown Title')

        if link in existing_links:
            continue  # Skip already downloaded videos

        if duration_sec <= max_duration_seconds:
            log_text.insert(tk.END, f"下载中: {title} | {link} ({duration_sec} 秒)\n")
            root.update()

            def progress_hook(d):
                if d['status'] == 'downloading':
                    percent = d.get('_percent_str', '').strip()
                    log_text.delete(1.0, tk.END)  # Clear previous log
                    log_text.insert(tk.END, f"正在下载: {title} - 下载进度: {percent}\n")
                    log_text.see(tk.END)
                    root.update()

            success = download_video(link, quality, progress_callback=progress_hook)
            if success:
                new_links.append(link)
                downloaded_count += 1
                log_text.insert(tk.END, f"成功下载: {title}\n")
            else:
                log_text.insert(tk.END, f"下载失败，跳过: {title}\n")
        else:
            log_text.insert(tk.END, f"时长超限，跳过: {title} | {link} ({duration_sec} 秒)\n")
        
        progress_bar['value'] = (downloaded_count / desired_download_count) * 100
        root.update()

    if new_links:
        append_links_to_excel(EXCEL_FILE, new_links)
        log_text.insert(tk.END, f"\n成功下载 {len(new_links)} 个视频，已写入 {EXCEL_FILE}\n")
    else:
        log_text.insert(tk.END, "\n无下载视频\n")

    progress_bar['value'] = 0

# ----------------- 启动按钮事件 -----------------
def start_download():
    keyword = entry_keyword.get().strip()
    if not keyword:
        messagebox.showerror("错误", "请输入搜索关键词")
        return

    try:
        desired_download_count = int(entry_count.get())
        max_duration_minutes = int(entry_duration.get())
    except ValueError:
        messagebox.showerror("错误", "请输入有效数字")
        return

    quality = quality_var.get()

    threading.Thread(target=download_thread, args=(keyword, desired_download_count, max_duration_minutes, quality)).start()

# ----------------- GUI 界面 -----------------
root = tk.Tk()
root.title("YouTube视频下载 (MP4)")
root.geometry("700x600")

frame = tk.Frame(root)
frame.pack(pady=10)

tk.Label(frame, text="搜索关键词:").grid(row=0, column=0)
entry_keyword = tk.Entry(frame, width=40)
entry_keyword.grid(row=0, column=1, padx=5)


tk.Label(frame, text="下载数量:").grid(row=1, column=0)
entry_count = tk.Entry(frame, width=10)
entry_count.insert(0, "2")
entry_count.grid(row=1, column=1, sticky="w", padx=5)


tk.Label(frame, text="时长限制(分钟):").grid(row=2, column=0)
entry_duration = tk.Entry(frame, width=10)
entry_duration.insert(0, "10")
entry_duration.grid(row=2, column=1, sticky="w", padx=5)

# 画质选择
tk.Label(frame, text="选择画质:").grid(row=3, column=0)
quality_var = tk.StringVar(value='480p')
quality_menu = ttk.Combobox(frame, textvariable=quality_var, values=['480p', '720p', '1080p'], width=8)
quality_menu.grid(row=3, column=1, sticky="w", padx=5)

start_button = tk.Button(root, text="开始下载", command=start_download)
start_button.pack(pady=10)

progress_bar = ttk.Progressbar(root, length=600, mode='determinate')
progress_bar.pack(pady=5)

log_text = tk.Text(root, height=25)
log_text.pack(fill="both", expand=True, padx=10, pady=10)

root.mainloop()
